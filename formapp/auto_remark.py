import numpy as np
import pandas as pd
from openpyxl import load_workbook
from formapp.models import reestr
from django.db.models import Q
from datetime import datetime
from formsite.settings import MEDIA_ROOT
import os, sys

def auto_import(file, reest):
    db = pd.read_excel(file, header=0, converters={'Номер замечания в проекте': str})
    closed = []
    returned = {}
    for i, row in db.iterrows():
        status_in_file = row['Статус'].replace("_x000D_", "\n")
        if status_in_file == "Устранено":
            closed.append(row['Номер замечания в проекте'].replace("_x000D_", "\n") + '[' + str(row['ID замечания']) + ']')
        elif status_in_file == "Повторно":
            returned[row['Номер замечания в проекте'].replace("_x000D_", "\n") + '[' + str(row['ID замечания']) + ']'] = row['Причина повторного направления'].replace("_x000D_", "\n")
        elif status_in_file in ["Ответ не предоставлен", "Устраняется"]:
            created_remarks = reestr.objects.filter((Q(num_remark=row['Номер замечания в проекте'].replace("_x000D_", "\n") + '[' + str(row['ID замечания']) + ']') & Q(actuality=True) & Q(reestrID=reest.id)))
            if len(created_remarks) == 0:
                reestr.objects.create(reestrID=reest,
                                      customer=reest.customer,
                                      project_dogovor=reest.project_dogovor,
                                      project_date_contract=reest.project_date_contract,
                                      project_name=reest.project_name,
                                      gip=reest.gip,
                                      project_reviewer=reest.project_reviewer,
                                      out_mail_num=reest.out_mail_num,
                                      out_mail_date=reest.out_mail_date,
                                      in_mail_num=reest.in_mail_num,
                                      in_mail_date=reest.in_mail_date,
                                      num_reestr=reest.num_reestr,
                                      num_remark=row['Номер замечания в проекте'].replace("_x000D_", "\n") + '[' + str(
                                          row['ID замечания']) + ']',
                                      designation_name=row['Ссылка на материалы'].replace("_x000D_", "\n"),
                                      section_name=row['Раздел документации'].replace("_x000D_", "\n"),
                                      remark_name=row['Вывод о несоответствии'].replace("_x000D_", "\n"),
                                      rational=row['Основание'].replace("_x000D_", "\n"),
                                      status="На заполнении ГИПом",
                                      remark_v=0,
                                      answer_remark="")
    closed_remarks = reestr.objects.filter((Q(num_remark__in=closed) & Q(actuality=True) & Q(reestrID=reest.id)))
    returned_remarks = reestr.objects.filter((Q(num_remark__in=returned.keys()) & Q(actuality=True) & Q(reestrID=reest.id) & (Q(status="На согласовании Рецензентом") | Q(status="Замечание снято"))))
    for r in closed_remarks:
        r.status = "Замечание снято"
        r.save(update_fields=['status'])
    for r in returned_remarks:
        r.actuality = False
        r.save(update_fields=['actuality'])
        d = datetime.now()
        reestr.objects.create(reestrID=r.reestrID,
                              customer=r.customer,
                              project_dogovor=r.project_dogovor,
                              project_date_contract=r.project_date_contract,
                              project_name=r.project_name,
                              gip=r.gip,
                              project_reviewer=r.project_reviewer,
                              out_mail_num=r.out_mail_num,
                              out_mail_date=r.out_mail_date,
                              in_mail_num=r.in_mail_num,
                              in_mail_date=r.in_mail_date,
                              num_reestr=r.num_reestr,
                              num_remark=r.num_remark,
                              designation_name=r.designation_name,
                              section_name=r.section_name,
                              remark_name=r.remark_name,
                              rational=r.rational,
                              status="На заполнении ГИПом",
                              remark_v=r.remark_v + 1,
                              answer_remark=r.answer_remark,
                              comment=r.comment + "\n" + returned[r.num_remark] + " (" + str(d.hour) + ':' + str(d.minute) +
                                                    ':' + str(d.second) + ' ' + str(d.day) + '.' + str(d.month) +
                                                    '.' + str(d.year) + ' ГГЭ)')

def auto_export(file, reest):
    try:
        file_path = os.path.join(MEDIA_ROOT, str(file))
        workbook = load_workbook(filename=file_path, )
        sheet = workbook.active
        db = pd.read_excel(file_path, header=0, converters={'Номер замечания в проекте': str})
        remark_ids = np.array(db.loc[:, 'ID замечания'])
        remark_nums = np.array(db.loc[:, 'Номер замечания в проекте'])
        # remark_text = np.array(db.loc[:, 'Вывод о несоответствии'])
        # remark_link = np.array(db.loc[:, 'Раздел документации'])
        lens = []
        for i in range(len(remark_nums)):
            #remarks = reestr.objects.filter((Q(actuality=True) & Q(status="Принято ГИПом") & Q(reestrID=reest.id) & Q(remark_name=remark_text[i].replace("_x000D_", "\n")) & Q(section_name=remark_link[i].replace("_x000D_", "\n"))))
            remarks = reestr.objects.filter((Q(actuality=True) & Q(status="Принято ГИПом") & Q(reestrID=reest.id) & Q(num_remark=remark_nums[i].replace("_x000D", "\n")+'['+remark_ids[i]+']')))
            lens.append(len(remarks))
            if len(remarks) > 0:
                str_answer = ''
                for r in remarks:
                    if r.num_remark[:r.num_remark.find('[')] == remark_nums[i]:
                        if r.answer_remark is not None:
                            str_answer += r.answer_remark + "\n\n"
                        if r.link_tech_name is not None:
                            str_answer += r.link_tech_name + "\n\n"
                        r.status = "На согласовании Рецензентом"
                        r.save(update_fields=['status'])
                sheet.cell(row=i + 2, column=15).value = str_answer
        workbook.save(file_path)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        print("Ошибка", e, exc_tb.tb_lineno)