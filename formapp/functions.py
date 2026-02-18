import os.path, sys
import smtplib
from email.header import Header
import xlsxwriter
from openpyxl import load_workbook
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from formapp.models import reestr, reestInfo, reviewers
from changelog.models import ChangeLog
from django.db.models import Q
from formsite.settings import MEDIA_ROOT
from email.mime.text import MIMEText
from django.contrib.auth.models import User, Group


def email_sender(mail_to, message, id):
    print(mail_to, message, id)
    # s = smtplib.SMTP('autodiscover.vnipipt.ru', 25)
    # # запуск TLS
    # # Авторизация
    # email = "crds@vnipipt.ru"
    # # Содержимое письма
    # subject = "Новая задача по реестру " + message
    # message = "Поступила новая задача по реестру № " + message + "\n\nhttp://crds.vnipipt.ru/reestr/" + str(id) + "/"
    # mime = MIMEText(message, 'plain', 'utf-8')
    # mime['Subject'] = Header(subject, 'utf-8')
    # # отправка письма
    # s.sendmail(email, mail_to, mime.as_string())
    # # прерывание сессии
    # s.quit()
    # print(email, mail_to, subject, message)


def dataFormat(s):
    months = {'января': '01', 'февраля': '02', 'марта': '03', 'апреля': '04', 'мая': '05', 'июня': '06', 'июля': '07',
              'августа': '08', 'сентября': '09', 'октября': '10', 'ноября': '11', 'декабря': '12'}
    i = 0
    day = ''
    while s[i] in '0123456789 ' and i < len(s):
        if s[i] != ' ':
            day += s[i]
        i += 1
    if len(day) < 2:
        day = '0' + day
    month = ''
    while s[i] != ' ':
        month += s[i]
        i += 1
    month = months[month]
    year = ''
    i += 1
    while s[i] in '0123456789':
        year += s[i]
        i += 1
    return day + '.' + month + '.' + year

def plannerDateFormat(s):
    months = {'Январь': '01', 'Февраль': '02', 'Март': '03', 'Апрель': '04', 'Май': '05', 'Июнь': '06',
              'Июль': '07',
              'Август': '08', 'Сентябрь': '09', 'Октябрь': '10', 'Ноябрь': '11', 'Декабрь': '12'}
    months_rev = {months[key]: key for key in months}
    if isinstance(s, str):
        j = 0
        day = ''
        while s[j] in '0123456789 ' and j < len(s):
            if s[j] != ' ':
                day += s[j]
            j += 1
        month = ''
        while s[j] != ' ':
            month += s[j]
            j += 1
        month = months[month]
        year = ''
        j += 1
        while s[j] in '0123456789':
            year += s[j]
            j += 1
        result = datetime.strptime(day + '.' + month + '.' + year, '%d.%m.%Y').date()
    else:
        result = ''
        if s.day < 10:
            result += '0'
        result += str(s.day) + ' '
        if s.month < 10:
            result += months_rev['0'+str(s.month)]
        else:
            result += months_rev[str(s.month)]
        result += ' ' + str(s.year)
    return result

def dateDBformat(d):
    if d is not None:
        if d.month < 10:
            month = '0' + str(d.month)
        else:
            month = str(d.month)
        return str(d.day) + '.' + month + '.' + str(d.year)
    else:
        return '     '


def shortName(s):
    if s.first_name != '(субподрядчик)':
        i = 0
        while s.first_name[i] != ' ':
            i += 1
        newS = s.last_name + ' ' + s.first_name[0] + '.' + s.first_name[i + 1] + '.'
        return newS
    else:
        return s.last_name


def select_dir():
    root = tk.Tk()
    root.withdraw()
    root.wm_attributes('-topmost', 1)
    return filedialog.askdirectory(parent=root)


def getHumanReadable(size):
    suffixes = [' Б', ' КБ', ' МБ', ' ГБ', ' ТБ']
    suffixIndex = 0
    while size > 1024:
        suffixIndex += 1  # increment the index of the suffix
        size = float('{:.2f}'.format(size / 1024.0))  # apply the division
    return str(size) + suffixes[suffixIndex]


def workDays(day, delay):
    k = 0
    while k < delay:
        day += timedelta(days=1)
        if day.weekday() < 5 and day not in govHolidays:
            k += 1
    return day
#учёт праздничных/рабочих дней календаря
govHolidays = [
    datetime.strptime('1.1.2025', '%d.%m.%Y').date(),
    datetime.strptime('2.1.2025', '%d.%m.%Y').date(),
    datetime.strptime('3.1.2025', '%d.%m.%Y').date(),
    datetime.strptime('6.1.2025', '%d.%m.%Y').date(),
    datetime.strptime('7.1.2025', '%d.%m.%Y').date(),
    datetime.strptime('8.1.2025', '%d.%m.%Y').date(),
    datetime.strptime('1.5.2025', '%d.%m.%Y').date(),
    datetime.strptime('2.5.2025', '%d.%m.%Y').date(),
    datetime.strptime('8.5.2025', '%d.%m.%Y').date(),
    datetime.strptime('9.5.2025', '%d.%m.%Y').date(),
    datetime.strptime('12.6.2025', '%d.%m.%Y').date(),
    datetime.strptime('13.6.2025', '%d.%m.%Y').date(),
    datetime.strptime('3.11.2025', '%d.%m.%Y').date(),
    datetime.strptime('4.11.2025', '%d.%m.%Y').date(),
    datetime.strptime('31.12.2025', '%d.%m.%Y').date(),
    datetime.strptime('1.1.2026', '%d.%m.%Y').date(),
    datetime.strptime('2.1.2026', '%d.%m.%Y').date(),
    datetime.strptime('5.1.2026', '%d.%m.%Y').date(),
    datetime.strptime('6.1.2026', '%d.%m.%Y').date(),
    datetime.strptime('7.1.2026', '%d.%m.%Y').date(),
    datetime.strptime('8.1.2026', '%d.%m.%Y').date(),
    datetime.strptime('9.1.2026', '%d.%m.%Y').date(),
    datetime.strptime('23.2.2026', '%d.%m.%Y').date(),
    datetime.strptime('9.3.2026', '%d.%m.%Y').date(),
    datetime.strptime('1.5.2026', '%d.%m.%Y').date(),
    datetime.strptime('11.5.2026', '%d.%m.%Y').date(),
    datetime.strptime('12.6.2026', '%d.%m.%Y').date(),
    datetime.strptime('4.11.2026', '%d.%m.%Y').date(),
    datetime.strptime('31.12.2026', '%d.%m.%Y').date(),
]
work_weekends = [datetime.strptime('1.11.2025', '%d.%m.%Y').date()]
def workDelay(day1, day2):
    day = day1
    delay = 0
    while day != day2:
        day += timedelta(days=1)
        if (day.weekday() < 5 and day not in govHolidays):# or day in work_weekends:
            delay += 1
    return delay

subcontracts = Group.objects.get(name="Субподрядчик").user_set.all()
def xlsxPlanner(file, reest):
    try:
        file_path = os.path.join(MEDIA_ROOT, str(file))
        workbook = load_workbook(filename=file_path, )
        sheet = workbook.active
        db = pd.read_excel(file_path, header=0, converters={'Окончание': str})
        department_list = np.array(db.loc[:, 'Выдает'])
        labour = np.array(db.loc[:, 'Трудозатраты'])
        start_date = np.array(db.loc[:, 'Начало'])
        final_date = np.array(db.loc[:, 'Окончание'])
        work_days = np.array(db.loc[:, 'Длительность'])
        for i in range(len(department_list)):
            start_date[i] = plannerDateFormat(start_date[i])
            labour[i] = float(labour[i][:-1].replace(',', '.'))
            if isinstance(department_list[i], str):
                brkt = department_list[i].find('(')
                if brkt > -1:
                    department_list[i] = department_list[i][:brkt]
        dep_uniq, dep_idx, dep_count = np.unique(department_list[1:], return_inverse=True, return_counts=True)
        proportion = np.full(len(department_list), 1.0)
        dep_uniq_idxs = []
        for i in range(len(dep_count)):
            idxs = np.where(dep_idx == i)
            idxs = idxs + np.full(len(idxs), 1)
            dep_uniq_idxs.append(idxs[0])
            if dep_count[i] > 1:
                sum = np.sum(labour[idxs][0])
                for j in idxs[0]:
                    proportion[j] = Decimal(str(labour[j]/sum)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
        dep_uniq = [dep_uniq, np.zeros(len(dep_uniq)), np.full(len(dep_uniq), datetime.strptime('1.1.1960', '%d.%m.%Y').date())]
        remarks = reestr.objects.filter((Q(reestrID=reest) & Q(actuality=True) & Q(department__in=department_list[1:]))).exclude(executor_name__in=subcontracts)
        remarks2 = reestr.objects.filter((Q(reestrID=reest) & Q(actuality=True) & Q(executor_name__in=subcontracts)))
        for i in remarks:
            if i.department != 'ОКиВД':
                idx = np.where(dep_uniq[0]==i.department)
                dep_uniq[1][idx[0][0]] += i.labor_costs_plan
                if dep_uniq[2][idx[0][0]] < i.answer_deadline_correct_plan:
                    dep_uniq[2][idx[0][0]] = i.answer_deadline_correct_plan
        for i in remarks2:
            idx = np.where(dep_uniq[0] == "Субподряд")
            dep_uniq[1][idx[0][0]] += i.labor_costs_plan
            if dep_uniq[2][idx[0][0]] < i.answer_deadline_correct_plan:
                dep_uniq[2][idx[0][0]] = i.answer_deadline_correct_plan
        for i in range(1, len(department_list)):
            if department_list[i] != 'ОКиВД':
                idx = np.where(dep_uniq[0] == department_list[i])
                labour[i] = float(Decimal(str(dep_uniq[1][idx[0][0]]*proportion[i])).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                if dep_uniq[2][idx[0][0]] == datetime.strptime('1.1.1960', '%d.%m.%Y').date():
                    final_date[i] = start_date[i]
                    work_days[i] = 0
                else:
                    final_date[i] = dep_uniq[2][idx[0][0]]
                    work_days[i] = workDelay(start_date[i], final_date[i])+1
                sheet.cell(row=i + 2, column=11).value = plannerDateFormat(final_date[i]) + ' 18:00' # окончание
                sheet.cell(row=i + 2, column=12).value = str(work_days[i]) + ' дней' # длительность
                sheet.cell(row=i + 2, column=13).value = str(labour[i]) + 'д' # трудозатраты
            else:
                final_date[i] = plannerDateFormat(final_date[i])
        labour[0] = float(Decimal(str(np.sum(labour[1:]))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
        final_date[0] = np.max(final_date[1:])
        work_days[0] = workDelay(start_date[0], final_date[0])+1
        sheet.cell(row=2, column=11).value = plannerDateFormat(final_date[0]) + ' 18:00'  # окончание
        sheet.cell(row=2, column=12).value = str(work_days[0]) + ' дней'  # длительность
        sheet.cell(row=2, column=13).value = str(labour[0]) + 'д'  # трудозатраты
        workbook.save(file_path)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        print("Ошибка", e, exc_tb.tb_lineno)

def xlsxGIPplannerCheckCreate(request, reest):
    remarks = reestr.objects.filter(reestrID=reest).order_by('department')
    departments = [r.department for r in remarks]
    departments = np.unique(np.array(departments))
    print(departments)
    remarks_structured = [[] for i in range(len(departments)+1)]
    d = 0
    for r in remarks:
        if r.executor_name in subcontracts:
            remarks_structured[-1].append(r)
        else:
            if r.department != departments[d]:
                d += 1
            remarks_structured[d].append(r)
    date = datetime.now()
    name = "\сверка_трудозатрат_" + str(reest.project_dogovor)[4:9] + str(reest.num_reestr) + "_" + str(
        date.hour) + '-' + str(date.minute) + '-' + str(date.second) + '_' + str(date.day) + '-' + str(
        date.month) + '-' + str(date.year) + '.xlsx'
    workbook = xlsxwriter.Workbook(MEDIA_ROOT + "\\Tables" + name)
    worksheet = workbook.add_worksheet()
    worksheet.hide_gridlines(2)
    cell_format2wrap = workbook.add_format({'bold': True,  # шапка таблицы с переносом
                                            'font_size': 14,
                                            'font_name': 'Times New Roman',
                                            'align': 'center',
                                            'valign': 'vcenter',
                                            'border': 1,
                                            'bg_color': '#DDEBF7',
                                            'text_wrap': True})
    cell_format4 = workbook.add_format({'font_size': 14,  # замечания
                                        'font_name': 'Times New Roman',
                                        'align': 'center',
                                        'valign': 'vcenter',
                                        'border': 1})
    cell_format4left = workbook.add_format({'font_size': 14,  # замечания по левому краю
                                            'font_name': 'Times New Roman',
                                            'align': 'left',
                                            'valign': 'vcenter',
                                            'border': 1,
                                            'text_wrap': True})
    worksheet.set_column('A:A', 3.89)
    worksheet.set_column('B:B', 3.57)
    worksheet.set_column('C:C', 3.57)
    worksheet.set_column('D:D', 15.77)
    worksheet.set_column('E:E', 3.68)
    worksheet.set_column('F:F', 3.51)
    worksheet.set_column('G:G', 3.68)
    worksheet.set_column('H:H', 3.51)
    worksheet.write(0, 0, "Подразделение", cell_format2wrap)
    worksheet.write(0, 1, "№ Замечания", cell_format2wrap)
    worksheet.write(0, 2, "Версия", cell_format2wrap)
    worksheet.write(0, 3, "Наименование замечания", cell_format2wrap)
    worksheet.write(0, 4, "Трудозатраты, дн. (на устранение замечания) (План)", cell_format2wrap)
    worksheet.write(0, 5, "Суммарные плановые трудозатраты, дн.", cell_format2wrap)
    worksheet.write(0, 6, "Трудозатраты, дн. (на устранение замечания) (Факт)", cell_format2wrap)
    worksheet.write(0, 7, "Суммарные фактические трудозатраты, дн.", cell_format2wrap)
    k = 1
    for i in range(len(remarks_structured)):
        sum1 = 0
        sum2 = 0
        buf_k = k
        for j in remarks_structured[i]:
            worksheet.write(k, 1, j.num_remark, cell_format4left)
            worksheet.write(k, 2, j.remark_v, cell_format4left)
            worksheet.write(k, 3, j.remark_name, cell_format4left)
            if j.labor_costs_plan is not None:
                worksheet.write(k, 4, j.labor_costs_plan, cell_format4)
                sum1 += j.labor_costs_plan
            else:
                worksheet.write(k, 4, '-', cell_format4)
            if j.labor_costs_fact is not None:
                worksheet.write(k, 6, j.labor_costs_fact, cell_format4)
                sum2 += j.labor_costs_fact
            else:
                worksheet.write(k, 6, '-', cell_format4)
            k += 1
        dep = ''
        if i < len(remarks_structured) - 1:
            dep = departments[i]
        elif i == len(remarks_structured) - 1 and len(remarks_structured[i]) > 0:
            dep = "Субподряд"
        elif i == len(remarks_structured) - 1 and len(remarks_structured[i]) == 0:
            break
        if len(remarks_structured[i]) > 1:
            worksheet.merge_range("A" + str(buf_k + 1) + ":A" + str(k), dep, cell_format4left)
            worksheet.merge_range("F" + str(buf_k + 1) + ":F" + str(k), str(sum1), cell_format4)
            worksheet.merge_range("H" + str(buf_k + 1) + ":H" + str(k), str(sum2), cell_format4)
        else:
            worksheet.write(buf_k, 0, departments[i], cell_format4left)
            worksheet.write(buf_k, 5, str(sum1), cell_format4)
            worksheet.write(buf_k, 7, str(sum2), cell_format4)
    # worksheet.autofit()
    workbook.close()
    return name

def xlslxCauseCreate(request, begin_date, end_date):
    name = "\aформулировки_коренных_причин_" + begin_date + '_' + end_date + '.xlsx'
    workbook = xlsxwriter.Workbook(MEDIA_ROOT + "\\Tables" + name)
    worksheet = workbook.add_worksheet()
    worksheet.hide_gridlines(2)
    cell_format2 = workbook.add_format({'bold': True,  # шапка таблицы
                                        'font_size': 20,
                                        'font_name': 'Times New Roman',
                                        'align': 'center',
                                        'valign': 'vcenter',
                                        'border': 1,
                                        'bg_color': '#DDEBF7'})
    cell_format4wrap = workbook.add_format({'font_size': 20,  # замечания c переносом
                                            'font_name': 'Times New Roman',
                                            'align': 'center',
                                            'valign': 'vcenter',
                                            'border': 1,
                                            'text_wrap': True})
    worksheet.set_column('A:A', 5.32)
    worksheet.set_column('B:C', 12.10)
    worksheet.set_column('E:E', 12.10)
    worksheet.set_column('D:D', 5.32)
    worksheet.write(0, 0, "Номер замечания", cell_format2)
    worksheet.write(0, 1, "Формулировка коренной причины", cell_format2)
    worksheet.write(0, 2, "Рецензент", cell_format2)
    worksheet.write(0, 3, "Подразделение", cell_format2)
    worksheet.write(0, 4, "Наименование замечания", cell_format2)
    remarks = reestr.objects.filter((Q(in_mail_date__gte=begin_date) & Q(in_mail_date__lte=end_date) & Q(root_cause_list__icontains='0.')))
    l = 1
    for i in remarks:
        remark_number = i.remark_index[4:i.remark_index.find("-")]+i.remark_index[i.remark_index.find("Д")+1:i.remark_index.rfind("_")]
        worksheet.write(l, 0, remark_number.replace("_", "-"), cell_format4wrap)
        worksheet.write(l, 1, i.root_cause_text, cell_format4wrap)
        worksheet.write(l, 2, reviewers.objects.get(id=i.project_reviewer.id).name, cell_format4wrap)
        worksheet.write(l, 3, i.department, cell_format4wrap)
        worksheet.write(l, 4, i.remark_name, cell_format4wrap)
        l += 1
    #worksheet.autofit()
    workbook.close()
    return name

def xlslxStatusCreate(request, remarks):
    date = datetime.now()
    r = reestr.objects.get(id=remarks[0])
    name = "\динамика_статусов_" + str(r.project_dogovor)[4:9] + str(r.num_reestr) + "_" + str(
        date.hour) + '-' + str(date.minute) + '-' + str(date.second) + '_' + str(date.day) + '-' + str(
        date.month) + '-' + str(date.year) + '.xlsx'
    workbook = xlsxwriter.Workbook(MEDIA_ROOT + "\\Tables" + name)
    worksheet = workbook.add_worksheet()
    worksheet.hide_gridlines(2)
    cell_format2 = workbook.add_format({'bold': True,  # шапка таблицы
                                        'font_size': 20,
                                        'font_name': 'Times New Roman',
                                        'align': 'center',
                                        'valign': 'vcenter',
                                        'border': 1,
                                        'bg_color': '#DDEBF7'})
    cell_format4wrap = workbook.add_format({'font_size': 20,  # замечания c переносом
                                            'font_name': 'Times New Roman',
                                            'align': 'center',
                                            'valign': 'vcenter',
                                            'border': 1,
                                            'text_wrap': True})
    worksheet.set_column('A:A', 5.32)
    worksheet.set_column('B:C', 12.10)
    worksheet.set_column('D:E', 5.32)
    worksheet.write(0, 0, "Номер замечания", cell_format2)
    worksheet.write(0, 1, "Статус", cell_format2)
    worksheet.write(0, 2, "Ответственный за этап", cell_format2)
    worksheet.write(0, 3, "Время наступления", cell_format2)
    worksheet.write(0, 4, "Время завершения", cell_format2)
    l = 1
    k = 1
    for i in remarks:
        remarkObj = reestr.objects.get(id=i)
        remark_number = remarkObj.remark_index[4:remarkObj.remark_index.find("-")] + remarkObj.remark_index[remarkObj.remark_index.find("Д") + 1:remarkObj.remark_index.rfind("_")]
        remark_number = remark_number.replace("_", "-")
        logs = ChangeLog.objects.filter((Q(model="Замечания") & Q(object_id=i))).order_by('changed')
        k = l
        for j in range(len(logs)):
            try:
                worksheet.write(l, 0, remark_number, cell_format4wrap)
                worksheet.write(l, 1, logs[j].data["Статус"], cell_format4wrap)
                respons = ""
                if logs[j].data["Статус"] in ["Формирование", "На заполнении ГИПом", "На согласовании ГИПом", "Согласовано ГИПом", "Принято ГИПом", "На доработке ГИПом", "Подготовка ответов ГИПом"]:
                    respons = remarkObj.gip.last_name + " " + remarkObj.gip.first_name
                elif logs[j].data["Статус"] in ["На заполнении руководителем", "На согласовании руководителем", "На доработке руководителем", "Подготовка ответов руководителем"]:
                    respons = remarkObj.responsibleTrouble_name.last_name + " " + remarkObj.responsibleTrouble_name.first_name
                elif logs[j].data["Статус"] in ["На заполнении исполнителем", "Подготовка ответов исполнителем", "На доработке исполнителем"]:
                    respons = remarkObj.executor_name.last_name + " " + remarkObj.executor_name.first_name
                elif logs[j].data["Статус"] in ["На согласовании Рецензентом", "Замечание снято"]:
                    respons = ""
                worksheet.write(l, 2, respons, cell_format4wrap)
                worksheet.write(l, 3, logs[j].changed.date().strftime('%Y-%m-%d'), cell_format4wrap)
                worksheet.write(l, 4, "", cell_format4wrap)
                if k != l:
                    worksheet.write(l-1, 4, logs[j].changed.date().strftime('%Y-%m-%d'), cell_format4wrap)
                    k += 1
                l += 1
            except Exception as e:
                print("Статус не изменялся", e)
    # worksheet.autofit()
    workbook.close()
    return name

def xlslxCreate(request, userRole='Администратор', actual=True):
    date = datetime.now()
    # path = select_dir()
    name = "\реестр_" + str(request.POST.get("project_dogovor"))[4:9] + str(request.POST.get("num_reestr")) + "_" + str(
        date.hour) + '-' + str(date.minute) + '-' + str(date.second) + '_' + str(date.day) + '-' + str(
        date.month) + '-' + str(date.year) + '.xlsx'
    if actual:
        if (userRole == 'Руководитель'):
            remarks = reestr.objects.filter((Q(reestrID=request.POST.get("id")) & (
                        Q(responsibleTrouble_name=request.user) | Q(executor_name=request.user)) & Q(actuality=True)))
        elif (userRole == 'Исполнитель'):
            remarks = reestr.objects.filter(
                (Q(reestrID=request.POST.get("id")) & Q(executor_name=request.user) & Q(actuality=True)))
        elif (userRole == 'ГИП'):
            remarks = reestr.objects.filter((Q(reestrID=request.POST.get("id")) & (
                        Q(responsibleTrouble_name=request.user) | Q(executor_name=request.user) | Q(
                    gip=request.user)) & Q(actuality=True)))
        else:
            remarks = reestr.objects.filter((Q(reestrID=request.POST.get("id")) & Q(actuality=True)))
    else:
        if (userRole == 'Руководитель'):
            remarks = reestr.objects.filter((Q(reestrID=request.POST.get("id")) & (
                        Q(responsibleTrouble_name=request.user) | Q(executor_name=request.user)) & Q(actuality=True)))
            remarks_history = reestr.objects.filter((Q(reestrID=request.POST.get("id")) & (
                        Q(responsibleTrouble_name=request.user) | Q(executor_name=request.user)) & (Q(actuality=False) & Q(status="На согласовании Рецензентом"))))
        elif (userRole == 'Исполнитель'):
            remarks = reestr.objects.filter((Q(reestrID=request.POST.get("id")) & Q(executor_name=request.user) & Q(actuality=True)))
            remarks_history = reestr.objects.filter((Q(reestrID=request.POST.get("id")) & Q(executor_name=request.user) & (Q(actuality=False) & Q(status="На согласовании Рецензентом"))))
        elif (userRole == 'ГИП'):
            remarks = reestr.objects.filter((Q(reestrID=request.POST.get("id")) & (Q(responsibleTrouble_name=request.user) | Q(executor_name=request.user) | Q(gip=request.user)) & Q(actuality=True)))
            remarks_history = reestr.objects.filter((Q(reestrID=request.POST.get("id")) & (Q(responsibleTrouble_name=request.user) | Q(executor_name=request.user) | Q(gip=request.user)) & (Q(actuality=False) & Q(status="На согласовании Рецензентом"))))
        else:
            remarks = reestr.objects.filter((Q(reestrID=request.POST.get("id")) & Q(actuality=True)))
            remarks_history = reestr.objects.filter((Q(reestrID=request.POST.get("id")) & Q(actuality=False) & Q(status="На согласовании Рецензентом")))
    workbook = xlsxwriter.Workbook(MEDIA_ROOT + "\\Tables" + name)
    worksheet = workbook.add_worksheet()
    worksheet.hide_gridlines(2)

    cell_format1 = workbook.add_format({'bold': True,  # заголовок
                                        'font_size': 28,
                                        'font_name': 'Times New Roman',
                                        'align': 'center',
                                        'valign': 'vcenter'})
    cell_format1wrap = workbook.add_format({'bold': True,  # заголовок с переносом
                                            'font_size': 28,
                                            'font_name': 'Times New Roman',
                                            'align': 'center',
                                            'valign': 'vcenter',
                                            'text_wrap': True})
    cell_format2 = workbook.add_format({'bold': True,  # шапка таблицы
                                        'font_size': 20,
                                        'font_name': 'Times New Roman',
                                        'align': 'center',
                                        'valign': 'vcenter',
                                        'border': 1,
                                        'bg_color': '#DDEBF7'})
    cell_format2wrap = workbook.add_format({'bold': True,  # шапка таблицы с переносом
                                            'font_size': 20,
                                            'font_name': 'Times New Roman',
                                            'align': 'center',
                                            'valign': 'vcenter',
                                            'border': 1,
                                            'bg_color': '#DDEBF7',
                                            'text_wrap': True})
    cell_format3 = workbook.add_format({'bold': True,  # шапка таблицы с номерами
                                        'italic': True,
                                        'font_size': 20,
                                        'font_name': 'Times New Roman',
                                        'align': 'center',
                                        'valign': 'vcenter',
                                        'border': 1})
    cell_format4 = workbook.add_format({'font_size': 20,  # замечания
                                        'font_name': 'Times New Roman',
                                        'align': 'center',
                                        'valign': 'vcenter',
                                        'border': 1})
    cell_format4left = workbook.add_format({'font_size': 20,  # замечания по левому краю
                                            'font_name': 'Times New Roman',
                                            'align': 'left',
                                            'valign': 'vcenter',
                                            'border': 1,
                                            'text_wrap': True})
    cell_format4wrap = workbook.add_format({'font_size': 20,  # замечания c переносом
                                            'font_name': 'Times New Roman',
                                            'align': 'center',
                                            'valign': 'vcenter',
                                            'border': 1,
                                            'text_wrap': True})
    cell_format5 = workbook.add_format({'font_size': 24,  # подписи
                                        'font_name': 'Times New Roman',
                                        'align': 'left',
                                        'valign': 'bottom'})
    # заголовок
    worksheet.merge_range("C3:T3", "Реестр выявленных несоответствий № " + request.POST.get("project_dogovor")[
                                                                           4:9] + request.POST.get("num_reestr"),
                          cell_format1)
    worksheet.merge_range("C4:T4", "Наименование проекта: Договор № " + request.POST.get("project_dogovor") + " от " +
                          dataFormat(request.POST.get("project_date_contract")) + ' "' + request.POST.get(
        "project_name") + '"', cell_format1wrap)
    worksheet.merge_range("C5:T5", "(Рецензент : " + request.POST.get("project_reviewer") + "; исх. № " +
                          request.POST.get("out_mail_num") + " от " + dataFormat(
        request.POST.get("out_mail_date")) + "; вх. № " +
                          request.POST.get("in_mail_num") + " от " + dataFormat(request.POST.get("in_mail_date")) + ")",
                          cell_format1)

    # шапка
    worksheet.set_column('A:A', 8.33)
    worksheet.merge_range("A7:A8", " № ", cell_format2)

    worksheet.set_column('B:B', 24.56)
    worksheet.merge_range("B7:B8", " № замечания ", cell_format2)

    worksheet.set_column('C:C', 96.22)
    worksheet.merge_range("C7:C8", "Наименование замечания", cell_format2)

    worksheet.set_column('D:D', 63.78)
    worksheet.merge_range("D7:D8", "Обоснование", cell_format2)

    worksheet.set_column('E:E', 24.78)
    worksheet.merge_range("E7:E8", "Обозначение раздела в проекте", cell_format2wrap)

    worksheet.set_column('F:F', 28.33)
    worksheet.merge_range("F7:F8", "Наименование раздела", cell_format2wrap)

    worksheet.set_column('G:G', 29.33)
    worksheet.merge_range("G7:G8", "Ответственный за устранение замечания (начальник подразделения)", cell_format2wrap)

    worksheet.set_column('H:H', 29.89)
    worksheet.merge_range("H7:H8", "Исполнитель, допустивший замечание", cell_format2wrap)

    worksheet.set_column('I:I', 29.89)
    worksheet.merge_range("I7:I8", "Исполнитель, ответственный за устранение замечания", cell_format2wrap)

    worksheet.set_column('J:O', 25.33)
    worksheet.merge_range("J7:K7", "Дата предоставления ответов на замечания", cell_format2wrap)

    worksheet.merge_range("L7:M7", "Срок внесения корректировок", cell_format2wrap)

    worksheet.merge_range("N7:O7", "Трудозатраты, дн. (на устранение замечания)", cell_format2wrap)

    for i in range(9, 15):
        if i % 2 == 0:
            worksheet.write(7, i, "Факт", cell_format2)
        else:
            worksheet.write(7, i, "План", cell_format2)

    worksheet.set_column('P:P', 63.67)
    worksheet.merge_range("P7:P8", "Комментарии", cell_format2)

    worksheet.set_column('Q:Q', 63.56)
    worksheet.merge_range("Q7:Q8", "Ответы на замечания", cell_format2)

    worksheet.set_column('R:R', 38.33)
    worksheet.merge_range("R7:R8",
                          "Ссылка в технической документации (том, книга, раздел, стр.лист) на внесённые изменения",
                          cell_format2wrap)

    worksheet.set_column('S:S', 37.22)
    worksheet.merge_range("S7:S8", "Отметка о снятии замечания, дата", cell_format2wrap)

    worksheet.set_column('T:U', 34.33)
    worksheet.merge_range("T7:T8", "Значимость замечания", cell_format2wrap)

    worksheet.merge_range("U7:U8", "Коренная причина", cell_format2)

    if (userRole == 'Наблюдатель'):
        worksheet.merge_range("V7:V8", "Статус", cell_format2)

    # нумерация столбцов
    nums = ['1', '1.1', '2', '2.1', '3', '4', '5', '6', '6.1', '7', '8', '9', '10', '10.1', '10.2', '10.3', '11', '12',
            '13', '14', '15']
    for i in range(len(nums)):
        worksheet.write(8, i, nums[i], cell_format3)
    if (userRole == 'Наблюдатель'):
        worksheet.write(8, len(nums), '16', cell_format3)
    # замечания
    boss = []
    employers = []
    k = 0
    if not actual:
        nums = []
        for i in remarks_history:
            nums.append(i.num_remark)
    for i in range(len(remarks)):
        if remarks[i].responsibleTrouble_name is not None:
            boss.append(remarks[i].responsibleTrouble_name.id)
        if remarks[i].executor_name is not None:
            employers.append(remarks[i].executor_name.id)
        worksheet.write(9 + i + k, 0, str(i + 1), cell_format4)

        worksheet.write(9 + i + k, 1, remarks[i].num_remark, cell_format4)

        worksheet.write(9 + i + k, 2, remarks[i].remark_name, cell_format4left)

        worksheet.write(9 + i + k, 3, remarks[i].rational, cell_format4left)

        worksheet.write(9 + i + k, 4, remarks[i].designation_name, cell_format4wrap)

        worksheet.write(9 + i + k, 5, remarks[i].section_name, cell_format4wrap)

        if remarks[i].responsibleTrouble_name is not None:
            worksheet.write(9 + i + k, 6, shortName(remarks[i].responsibleTrouble_name), cell_format4)

        else:
            worksheet.write(9 + i + k, 6, '', cell_format4)

        if remarks[i].executor_fail_name is not None:
            if remarks[i].executor_fail_name.username != "emptyUSER" and remarks[
                i].executor_fail_name.username != "customer":
                worksheet.write(9 + i + k, 7, shortName(remarks[i].executor_fail_name), cell_format4)
            elif remarks[i].executor_fail_name.username == "emptyUSER":
                worksheet.write(9 + i + k, 7, remarks[i].executor_fail_text, cell_format4)
            elif remarks[i].executor_fail_name.username == "customer":
                worksheet.write(9 + i + k, 7, remarks[i].executor_fail_text, cell_format4)
        else:
            worksheet.write(9 + i + k, 7, '', cell_format4)

        if remarks[i].executor_name is not None:
            worksheet.write(9 + i + k, 8, shortName(remarks[i].executor_name), cell_format4)
        else:
            worksheet.write(9 + i + k, 8, '', cell_format4)
        worksheet.write(9 + i + k, 9, dateDBformat(remarks[i].answer_date_plan), cell_format4)

        worksheet.write(9 + i + k, 10, dateDBformat(remarks[i].answer_date_fact), cell_format4)

        worksheet.write(9 + i + k, 11, dateDBformat(remarks[i].answer_deadline_correct_plan), cell_format4)

        worksheet.write(9 + i + k, 12, dateDBformat(remarks[i].answer_deadline_correct_fact), cell_format4)

        worksheet.write(9 + i + k, 13, remarks[i].labor_costs_plan, cell_format4)

        worksheet.write(9 + i + k, 14, remarks[i].labor_costs_fact, cell_format4)

        comment_text = remarks[i].comment
        if remarks[i].root_cause_text:
            comment_text += "\n" + "Предложенная формулировка коренной причины: " + remarks[i].root_cause_text
        if remarks[i].root_cause_list is not None:
            comment_buf = ""
            if any(x in remarks[i].root_cause_list for x in ['1.3.1.', '1.3.2.', '3.1.1.', '3.1.2.']):
                comment_buf += "Конкретные пункты НД, "
            if '2.2.1.' in remarks[i].root_cause_list:
                comment_buf += "Конкретное ПО или оборудование, "
            if '3.1.5.' in remarks[i].root_cause_list:
                comment_buf += "Конкретный субподрядчик, "
            if '3.1.6.' in remarks[i].root_cause_list:
                comment_buf += "Конкретный исполнитель, "
            if any(x in remarks[i].root_cause_list for x in ['3.2.1.', '3.2.3.', '3.3.1.']):
                comment_buf += "Конкретные подразделения, "
            if any(x in remarks[i].root_cause_list for x in ['4.1.1.', '4.2.1.']):
                comment_buf += "Конкретные факты и Руководители, "
            if any(x in remarks[i].root_cause_list for x in ['4.1.2.', '4.2.2.']):
                comment_buf += "Конкретные факты, Руководители и конкретные распоряжения, "
            if '4.3.1.' in remarks[i].root_cause_list:
                comment_buf += "Конкретные Руководители, "
            if any(x in remarks[i].root_cause_list for x in ['4.4.1.', '4.5.1.', '4.5.2.']):
                comment_buf += "Конкретные факты, "
            if '4.6.1.' in remarks[i].root_cause_list:
                comment_buf += "Конкретные Распоряжения и документы, "
            if '5.3.1.' in remarks[i].root_cause_list:
                comment_buf += "Конкретные Руководители (нормоконтролёры), "
            if '5.3.2.' in remarks[i].root_cause_list:
                comment_buf += "Конкретные Руководители и исполнители (нормоконтролёры), "
            if comment_buf != "":
                comment_text += "\n" + comment_buf[:-2] + ": " + str(remarks[i].root_cause_comment)
        # if remarks[i].importance3:
        #     comment_text += "\n" + "Требуется корректировка смежных разделов (глав, томов, др. док.): " + str(remarks[i].imp3_comment)
        # if remarks[i].importance4:
        #     comment_text += "\n" + "Требуется выдача задания смежным подразделениям (без учёта ОКиВД) и (или) получение исходных данных от них: " + str(remarks[i].imp4_comment)
        # if remarks[i].importance7:
        #     comment_text += "\n" + "Требуется получение дополнительных исходных данных от Заказчика: " + str(remarks[i].imp7_comment)
        worksheet.write(9 + i + k, 15, comment_text, cell_format4left)

        worksheet.write(9 + i + k, 16, remarks[i].answer_remark, cell_format4left)

        worksheet.write(9 + i + k, 17, remarks[i].link_tech_name, cell_format4wrap)

        cancel = ''
        if remarks[i].cancel_remark is not None and remarks[i].cancel_remark != '':
            cancel = remarks[i].cancel_remark + " от " + dateDBformat(remarks[i].cancel_remark_date)
        worksheet.write(9 + i + k, 18, cancel, cell_format4)

        worksheet.write(9 + i + k, 19, remarks[i].total_importance, cell_format4)

        worksheet.write(9 + i + k, 20, remarks[i].root_cause_list, cell_format4)

        if (userRole == 'Наблюдатель'):
            worksheet.write(9 + i + k, 21, remarks[i].status, cell_format4)

        if not actual and (remarks[i].num_remark in nums):
            v = 1
            for j in remarks_history:
                if j.num_remark == remarks[i].num_remark:
                    k += 1
                    worksheet.write(9 + i + k, 0, str(i + 1)+'.'+str(v), cell_format4)

                    worksheet.write(9 + i + k, 1, j.num_remark, cell_format4)

                    worksheet.write(9 + i + k, 2, j.remark_name, cell_format4left)

                    worksheet.write(9 + i + k, 3, j.rational, cell_format4left)

                    worksheet.write(9 + i + k, 4, j.designation_name, cell_format4wrap)

                    worksheet.write(9 + i + k, 5, j.section_name, cell_format4wrap)

                    if j.responsibleTrouble_name is not None:
                        worksheet.write(9 + i + k, 6, shortName(j.responsibleTrouble_name), cell_format4)

                    else:
                        worksheet.write(9 + i + k, 6, '', cell_format4)

                    if j.executor_fail_name is not None:
                        if j.executor_fail_name.username != "emptyUSER" and j.executor_fail_name.username != "customer":
                            worksheet.write(9 + i + k, 7, shortName(j.executor_fail_name), cell_format4)
                        elif j.executor_fail_name.username == "emptyUSER":
                            worksheet.write(9 + i + k, 7, j.executor_fail_text, cell_format4)
                        elif j.executor_fail_name.username == "customer":
                            worksheet.write(9 + i + k, 7, j.executor_fail_text, cell_format4)
                    else:
                        worksheet.write(9 + i + k, 7, '', cell_format4)

                    if j.executor_name is not None:
                        worksheet.write(9 + i + k, 8, shortName(j.executor_name), cell_format4)
                    else:
                        worksheet.write(9 + i + k, 8, '', cell_format4)
                    worksheet.write(9 + i + k, 9, dateDBformat(j.answer_date_plan), cell_format4)

                    worksheet.write(9 + i + k, 10, dateDBformat(j.answer_date_fact), cell_format4)

                    worksheet.write(9 + i + k, 11, dateDBformat(j.answer_deadline_correct_plan), cell_format4)

                    worksheet.write(9 + i + k, 12, dateDBformat(j.answer_deadline_correct_fact), cell_format4)

                    worksheet.write(9 + i + k, 13, j.labor_costs_plan, cell_format4)

                    worksheet.write(9 + i + k, 14, j.labor_costs_fact, cell_format4)

                    comment_text = remarks[i].comment
                    if j.root_cause_text:
                        comment_text += "\n" + "Предложенная формулировка коренной причины: " + j.root_cause_text
                    if j.root_cause_list is not None:
                        comment_buf = ""
                        if any(x in j.root_cause_list for x in ['1.3.1.', '1.3.2.', '3.1.1.', '3.1.2.']):
                            comment_buf += "Конкретные пункты НД, "
                        if '2.2.1.' in j.root_cause_list:
                            comment_buf += "Конкретное ПО или оборудование, "
                        if '3.1.5.' in j.root_cause_list:
                            comment_buf += "Конкретный субподрядчик, "
                        if '3.1.6.' in j.root_cause_list:
                            comment_buf += "Конкретный исполнитель, "
                        if any(x in j.root_cause_list for x in ['3.2.1.', '3.2.3.', '3.3.1.']):
                            comment_buf += "Конкретные подразделения, "
                        if any(x in j.root_cause_list for x in ['4.1.1.', '4.2.1.']):
                            comment_buf += "Конкретные факты и Руководители, "
                        if any(x in j.root_cause_list for x in ['4.1.2.', '4.2.2.']):
                            comment_buf += "Конкретные факты, Руководители и конкретные распоряжения, "
                        if '4.3.1.' in j.root_cause_list:
                            comment_buf += "Конкретные Руководители, "
                        if any(x in j.root_cause_list for x in ['4.4.1.', '4.5.1.', '4.5.2.']):
                            comment_buf += "Конкретные факты, "
                        if '4.6.1.' in j.root_cause_list:
                            comment_buf += "Конкретные Распоряжения и документы, "
                        if '5.3.1.' in j.root_cause_list:
                            comment_buf += "Конкретные Руководители (нормоконтролёры), "
                        if '5.3.2.' in j.root_cause_list:
                            comment_buf += "Конкретные Руководители и исполнители (нормоконтролёры), "
                        if comment_buf != "":
                            comment_text += "\n" + comment_buf[:-2] + ": " + str(j.root_cause_comment)
                    if j.importance3:
                        comment_text += "\n" + "Требуется корректировка смежных разделов (глав, томов, др. док.): " + str(j.imp3_comment)
                    if j.importance4:
                        comment_text += "\n" + "Требуется выдача задания смежным подразделениям (без учёта ОКиВД) и (или) получение исходных данных от них: " + str(j.imp4_comment)
                    if j.importance7:
                        comment_text += "\n" + "Требуется получение дополнительных исходных данных от Заказчика: " + str(j.imp7_comment)
                    worksheet.write(9 + i + k, 15, comment_text, cell_format4left)

                    worksheet.write(9 + i + k, 16, j.answer_remark, cell_format4left)

                    worksheet.write(9 + i + k, 17, j.link_tech_name, cell_format4wrap)

                    cancel = ''
                    if j.cancel_remark is not None and j.cancel_remark != '':
                        cancel = j.cancel_remark + " от " + dateDBformat(j.cancel_remark_date)
                    worksheet.write(9 + i + k, 18, cancel, cell_format4)

                    worksheet.write(9 + i + k, 19, j.total_importance, cell_format4)

                    worksheet.write(9 + i + k, 20, j.root_cause_list, cell_format4)

                    if (userRole == 'Наблюдатель'):
                        worksheet.write(9 + i + k, 21, j.status, cell_format4)


    if actual:
        k = 12 + len(remarks)
    else:
        k = 12 + len(remarks) + len(remarks_history)
    worksheet.write(k, 2, '"ЗНАЧИМОСТЬ ЗАМЕЧАНИЙ И КОРЕННЫЕ ПРИЧИНЫ СОГЛАСОВАЛ":', cell_format5)

    worksheet.write(k, 7, 'Заказчик: ' + request.POST.get("customer"), cell_format5)

    worksheet.write(k, 17, '"СНЯТИЕ ЗАМЕЧАНИЙ ПОДТВЕРЖДАЮ":', cell_format5)

    k += 2
    worksheet.write(k, 2, "Главный инженер проекта " + shortName(
        reestInfo.objects.get(id=request.POST.get("id")).gip) + " ___________(подпись) __________(дата)", cell_format5)

    worksheet.write(k, 17, "Главный инженер проекта " + shortName(
        reestInfo.objects.get(id=request.POST.get("id")).gip) + " ___________(подпись) __________(дата)", cell_format5)

    k += 3
    worksheet.write(k, 2, '"ЗНАЧИМОСТЬ ЗАМЕЧАНИЙ  И КОРЕННЫЕ ПРИЧИНЫ ОПРЕДЕЛИЛ":', cell_format5)

    k += 1
    worksheet.write(k, 17, '"ОТВЕТЫ НА  ЗАМЕЧАНИЯ ПРЕДСТАВИЛ":', cell_format5)

    k += 1
    boss = np.unique(np.array(boss))
    for i in range(len(boss)):
        b = User.objects.get(id=boss[i])
        worksheet.write(k, 2, 'Начальник отдела ' + shortName(b) + ' ___________(подпись) __________(дата)',
                        cell_format5)
        k += 1
        worksheet.write(k, 17, 'Начальник отдела ' + shortName(b) + ' ___________(подпись) __________(дата)',
                        cell_format5)
        k += 1
    worksheet.write(k, 2, '"С ОЦЕНКОЙ  ЗНАЧИМОСТИ ЗАМЕЧАНИЙ ОЗНАКОМЛЕН":', cell_format5)
    k += 2
    worksheet.write(k, 2, 'ИСПОЛНИТЕЛИ:', cell_format5)
    employers = np.unique(np.array(employers))
    for i in range(len(employers)):
        k += 1
        e = User.objects.get(id=employers[i])
        if e.username != "customer" and e.username != "emptyUSER":
            worksheet.write(k, 2, shortName(e) + ' ___________(подпись) __________(дата)', cell_format5)
    worksheet.autofit()
    workbook.close()
    return name


